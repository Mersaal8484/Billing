import io
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from odoo.exceptions import UserError, ValidationError
from odoo.tests.common import TransactionCase


class TestMigrationTemplateContract(TransactionCase):
    """Regression tests for the synchronized model/wizard/XLSX contract."""

    @classmethod
    def _template_path(cls, name):
        return Path(__file__).parents[1] / 'static' / 'src' / name

    def test_actual_templates_have_supported_contract(self):
        wizard = self.env['utility.migration.import.wizard'].new({})
        expected = {
            'Migration_Template.xlsx': ('customer_number', 'meter_number', 'opening_reading', 'phase', 'meter_model_code'),
            'Feeder_Migration_Template.xlsx': ('feeder_code', 'feeder_name', 'meter_number', 'current_reading', 'legacy_analytic_id', 'opening_reading'),
            'Transformer_Migration_Template.xlsx': ('reference', 'transformer_code', 'transformer_name', 'meter_number', 'current_reading', 'opening_reading', 'legacy_analytic_id'),
        }
        for filename, fields in expected.items():
            workbook = load_workbook(self._template_path(filename), data_only=True)
            self.assertIn('بيانات التهيئة', workbook.sheetnames)
            instructions = workbook['تعليمات الاستيراد']
            self.assertEqual(instructions['B3'].value, 2)
            headers = [wizard._normalize_header(c.value) for c in workbook['بيانات التهيئة'][4]]
            aliases = {wizard._normalize_header(alias): field for field, values in wizard.ALIASES.items() for alias in values}
            mapped = [aliases.get(header) for header in headers]
            for field in fields:
                self.assertIn(field, mapped, '%s is missing from %s' % (field, filename))
            counts = Counter(field for field in mapped if field)
            for field in fields:
                self.assertEqual(counts[field], 1, '%s is duplicated in %s' % (field, filename))

    def test_strict_parsers_distinguish_invalid_values(self):
        wizard = self.env['utility.migration.import.wizard'].new({})
        self.assertEqual(wizard.parse_float(0, 'reading', 8), 0.0)
        self.assertIsNone(wizard.parse_float(None, 'reading', 8))
        self.assertEqual(wizard.parse_phase('three', 8), 'three')
        with self.assertRaises(ValidationError):
            wizard.parse_float('15O.5', 'reading', 8)
        with self.assertRaises(ValidationError):
            wizard.parse_bool('YESS', field_name='is_active', row=8)
        with self.assertRaises(ValidationError):
            wizard.parse_phase('2', 8)
        with self.assertRaises(ValidationError):
            wizard._parse_multiplier(0, 8, 'meter_multiplier')

    def test_in_memory_workbook_missing_header_is_explicit(self):
        wizard = self.env['utility.migration.import.wizard'].new({'import_type': 'customer'})
        workbook = load_workbook(self._template_path('Migration_Template.xlsx'))
        sheet = workbook['بيانات التهيئة']
        sheet['D4'] = 'رقم قديم بدون هوية'
        sheet['D5'] = 'x'
        stream = io.BytesIO()
        workbook.save(stream)
        with self.assertRaises(UserError):
            wizard._read_contract(workbook)
