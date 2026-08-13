import base64
import binascii
import io
import re
import zipfile

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:  # pragma: no cover - reported to the user by the action
    openpyxl = None


class UtilityMigrationImportWizard(models.TransientModel):
    _name = 'utility.migration.import.wizard'
    _description = 'معالج استيراد بيانات التهيئة والميجريشن'

    TEMPLATE_VERSION = 2
    CONTRACTS = {
        'customer': {
            'required': ('name', 'customer_number', 'meter_number'),
            'fields': (
                'name', 'mobile', 'national_id', 'customer_number',
                'subscriber_no', 'char_code', 'is_active', 'legacy_region',
                'legacy_area', 'legacy_category', 'legacy_subscriber_type',
                'legacy_contract', 'meter_number', 'meter_reading',
                'opening_reading', 'previous_balance', 'current_balance',
                'phase', 'is_private_transformer', 'owner_reference',
                'meter_model_code',
            ),
        },
        'feeder': {
            'required': ('feeder_code', 'feeder_name', 'meter_number'),
            'fields': (
                'legacy_region', 'legacy_area', 'is_active', 'feeder_code',
                'feeder_name', 'meter_number', 'meter_multiplier',
                'current_reading', 'is_calculation_cell', 'description',
                'legacy_analytic_id', 'opening_reading',
            ),
        },
        'transformer': {
            'required': ('transformer_code', 'transformer_name', 'meter_number'),
            'fields': (
                'legacy_region', 'legacy_area', 'is_active', 'transformer_code',
                'transformer_name', 'meter_number', 'meter_multiplier',
                'current_reading', 'total_consumption', 'image_status',
                'cell_meter_number', 'cell_meter_multiplier', 'reference',
                'description', 'opening_reading', 'legacy_analytic_id',
            ),
        },
    }
    ALIASES = {
        'name': ('name', 'الاسم', 'الاسم *', 'اسم العرض / الخلية', 'اسم العرض / المحول'),
        'mobile': ('mobile', 'الموبايل', 'الموبايل *'),
        'national_id': ('national_id', 'الرقم الوطني'),
        'customer_number': ('customer_number', 'رقم المشترك', 'رقم المشترك *'),
        'subscriber_no': ('subscriber_no', 'الرقم الجديد'),
        'char_code': ('char_code', 'رقم الحرف'),
        'is_active': ('is_active', 'هل فعال؟', 'هل فعال؟ *'),
        'legacy_region': ('legacy_region', 'رمز المنطقة'),
        'legacy_area': ('legacy_area', 'رمز الفرع'),
        'legacy_category': ('legacy_category', 'رمز الفئة'),
        'legacy_subscriber_type': ('legacy_subscriber_type', 'رمز نوع المشترك'),
        'legacy_contract': ('legacy_contract', 'رمز قالب العقد', 'رمز قالب العقد *'),
        'meter_number': ('meter_number', 'رقم العداد', 'رقم العداد *', 'رقم عداد رصد المحول', 'رقم العداد (عداد الفيدر)'),
        'meter_reading': ('meter_reading', 'قراءة العداد في النظام', 'قراءة العداد في النظام القديم'),
        'opening_reading': ('opening_reading', 'قراءة الافتتاح', 'قراءة بداية الاشتراك', 'قراءة عند تفعيل العقد', 'القراءة عند تفعيل العقد'),
        'previous_balance': ('previous_balance', 'الرصيد السابق (الخط الساخن)'),
        'current_balance': ('current_balance', 'الرصيد الحالي (الافتتاحي)'),
        'phase': ('phase', 'نوع الفاز', 'نوع الفاز (single/three)', 'الطور'),
        'is_private_transformer': ('is_private_transformer', 'محول خاص?', 'محول خاص؟ (نعم/لا)'),
        'owner_reference': ('owner_reference', 'مرجع المالك القديم'),
        'meter_model_code': ('meter_model_code', 'رمز موديل العداد', 'meter model code'),
        'feeder_code': ('feeder_code', 'رمز الفيدر / الخلية *', 'رمز الفيدر / الحساب التحليلي'),
        'feeder_name': ('feeder_name', 'اسم الفيدر / الخلية *', 'اسم الفيدر / الخلية'),
        'meter_multiplier': ('meter_multiplier', 'معامل الضرب للعداد'),
        'current_reading': ('current_reading', 'القراءة الحالية'),
        'is_calculation_cell': ('is_calculation_cell', 'خلية إحتساب'),
        'description': ('description', 'الوصف'),
        'legacy_analytic_id': ('legacy_analytic_id', 'معرف الحساب التحليلي'),
        'transformer_code': ('transformer_code', 'رمز المحول *', 'رمز المحول / الحساب التحليلي'),
        'transformer_name': ('transformer_name', 'اسم المحول *', 'اسم المحول / الشريك'),
        'total_consumption': ('total_consumption', 'إجمالي الاستهلاك'),
        'image_status': ('image_status', 'حالة الصورة'),
        'cell_meter_number': ('cell_meter_number', 'رقم عداد الخلية المغذية', 'الخلية / رقم العداد'),
        'cell_meter_multiplier': ('cell_meter_multiplier', 'معامل ضرب عداد الخلية', 'الخلية / معامل الضرب للعداد'),
        'reference': ('reference', 'المرجع'),
    }

    import_type = fields.Selection([
        ('customer', 'تهيئة بيانات المشتركين'),
        ('feeder', 'تهيئة بيانات الفيدرات / الخلايا'),
        ('transformer', 'تهيئة بيانات المحولات'),
    ], string='نوع البيانات المراد استيرادها', default='customer', required=True)
    import_file = fields.Binary(string='ملف الإكسل', required=True)
    file_name = fields.Char(string='اسم الملف')

    def action_download_customer_template(self):
        return self._download('Migration_Template.xlsx')

    def action_download_feeder_template(self):
        return self._download('Feeder_Migration_Template.xlsx')

    def action_download_transformer_template(self):
        return self._download('Transformer_Migration_Template.xlsx')

    def _download(self, filename):
        return {'type': 'ir.actions.act_url', 'url': '/utility_core/static/src/%s' % filename, 'target': 'new'}

    @staticmethod
    def _has_cell_value(value):
        return value is not None and str(value).strip() != ''

    @staticmethod
    def _normalize_header(value):
        value = str(value or '').strip().lower()
        value = value.replace('؟', '?').replace('*', '')
        return re.sub(r'[^\w\u0600-\u06ff]+', '_', value).strip('_')

    def _error(self, code, row, field, value):
        return ValidationError(_('%s\nTemplate: %s\nRow: %s\nField: %s\nValue: %r') % (
            code, self.import_type, row, field, value))

    def parse_float(self, value, field_name='number', row=None):
        if not self._has_cell_value(value):
            return None
        try:
            result = float(value)
        except (ValueError, TypeError):
            raise self._error('INVALID_NUMERIC_VALUE', row or '?', field_name, value)
        if result < 0:
            raise self._error('INVALID_READING_VALUE', row or '?', field_name, value)
        return result

    def parse_int(self, value, field_name='integer', row=None):
        result = self.parse_float(value, field_name, row)
        if result is None:
            return None
        if not result.is_integer():
            raise self._error('INVALID_INTEGER_VALUE', row or '?', field_name, value)
        return int(result)

    def parse_bool(self, value, default=True, field_name='boolean', row=None):
        if not self._has_cell_value(value):
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)) and value in (0, 1):
            return bool(value)
        normalized = str(value).strip().lower()
        if normalized in ('false', '0', 'no', 'n', 'لا', 'خطأ', 'f'):
            return False
        if normalized in ('true', '1', 'yes', 'y', 'نعم', 'صح', 't'):
            return True
        raise self._error('INVALID_BOOLEAN_VALUE', row or '?', field_name, value)

    def parse_phase(self, value, row=None):
        if not self._has_cell_value(value):
            return None
        normalized = str(value).strip().lower()
        if normalized in ('1', '1 phase', 'single', 'single phase', 'أحادي', 'فاز واحد'):
            return 'single'
        if normalized in ('3', '3 phase', 'three', 'three phase', 'ثلاثي', 'ثلاثة فاز'):
            return 'three'
        raise self._error('INVALID_METER_PHASE', row or '?', 'phase', value)

    def _metadata_version(self, workbook):
        for sheet in workbook.worksheets:
            if self._normalize_header(sheet.title) in ('تعليمات_الاستيراد', 'instructions'):
                for row in sheet.iter_rows(values_only=True):
                    values = [str(v or '').strip().lower() for v in row]
                    if any(v in ('template version', 'إصدار القالب', 'template_version') for v in values):
                        for value in row:
                            if str(value or '').strip().isdigit():
                                return int(value)
        return 1  # immediately previous templates had no metadata

    def _read_contract(self, workbook):
        if self._metadata_version(workbook) not in (1, self.TEMPLATE_VERSION):
            raise UserError('UNSUPPORTED_MIGRATION_TEMPLATE_VERSION')
        contract = self.CONTRACTS[self.import_type]
        aliases = {}
        for field_name, names in self.ALIASES.items():
            for alias in names:
                aliases[self._normalize_header(alias)] = field_name
        data_sheet = next((s for s in workbook.worksheets if s.title != 'تعليمات الاستيراد'), None)
        if not data_sheet:
            raise UserError('MIGRATION_TEMPLATE_NO_DATA_SHEET')
        for header_row in range(1, min(data_sheet.max_row, 20) + 1):
            header_map = {}
            duplicates = []
            for index, cell in enumerate(data_sheet[header_row], start=1):
                field_name = aliases.get(self._normalize_header(cell.value))
                if field_name:
                    if field_name in header_map:
                        duplicates.append(field_name)
                    header_map[field_name] = index
            if header_map and len(set(contract['required']) & set(header_map)) == len(contract['required']):
                if duplicates:
                    raise UserError('DUPLICATE_NORMALIZED_HEADER: %s' % ', '.join(duplicates))
                missing = [f for f in contract['required'] if f not in header_map]
                if missing:
                    raise UserError('MISSING_REQUIRED_HEADER: %s' % ', '.join(missing))
                return data_sheet, header_row, header_map
        raise UserError('MISSING_REQUIRED_HEADER: %s' % ', '.join(contract['required']))

    def _cell(self, row, header_map, field_name):
        index = header_map.get(field_name)
        return row[index - 1] if index and index <= len(row) else None

    def _rows(self, sheet, header_row, header_map):
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(self._has_cell_value(value) for value in row):
                continue
            yield row_number, row

    def action_import_file(self):
        if not openpyxl:
            raise UserError(_('مكتبة openpyxl غير مثبتة.'))
        if not self.file_name or not self.file_name.lower().endswith('.xlsx'):
            raise UserError(_('يجب أن يكون الملف بصيغة .xlsx فقط.'))
        try:
            file_content = base64.b64decode(self.import_file, validate=True)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=False)
            sheet, header_row, header_map = self._read_contract(workbook)
        except (binascii.Error, ValueError, TypeError):
            raise UserError(_('INVALID_BASE64_FILE: ملف الإكسل المشفر غير صالح.'))
        except (OSError, KeyError, zipfile.BadZipFile, RuntimeError, openpyxl.utils.exceptions.InvalidFileException):
            raise UserError(_('INVALID_XLSX_FILE: الملف تالف أو محمي بكلمة مرور أو ليس ملف XLSX صالحًا.'))
        if self.import_type == 'customer':
            return self._import_customers(sheet, header_row, header_map)
        if self.import_type == 'feeder':
            return self._import_feeders(sheet, header_row, header_map)
        return self._import_transformers(sheet, header_row, header_map)

    def _identity_guard(self, seen, identity, row, code):
        if identity in seen:
            raise self._error(code, '%s and %s' % (seen[identity], row), 'identity', identity)
        seen[identity] = row

    def _import_customers(self, sheet, header_row, header_map):
        model = self.env['utility.migration.customer']
        records = model
        seen = {}
        company_id = self.env.company.id
        for row_number, row in self._rows(sheet, header_row, header_map):
            name = str(self._cell(row, header_map, 'name') or '').strip()
            number = str(self._cell(row, header_map, 'customer_number') or '').strip()
            if not name and not number:
                continue
            if not number:
                raise self._error('MISSING_CUSTOMER_NUMBER', row_number, 'customer_number', number)
            self._identity_guard(seen, number, row_number, 'DUPLICATE_CUSTOMER_NUMBER_IN_FILE')
            values = {
                'name': name,
                'mobile': str(self._cell(row, header_map, 'mobile') or '').strip(),
                'national_id': str(self._cell(row, header_map, 'national_id') or '').strip(),
                'customer_number': number,
                'subscriber_no': str(self._cell(row, header_map, 'subscriber_no') or '').strip(),
                'char_code': str(self._cell(row, header_map, 'char_code') or '').strip(),
                'is_active': self.parse_bool(self._cell(row, header_map, 'is_active'), True, 'is_active', row_number),
                'legacy_region': str(self._cell(row, header_map, 'legacy_region') or '').strip(),
                'legacy_area': str(self._cell(row, header_map, 'legacy_area') or '').strip(),
                'legacy_category': str(self._cell(row, header_map, 'legacy_category') or '').strip(),
                'legacy_subscriber_type': str(self._cell(row, header_map, 'legacy_subscriber_type') or '').strip(),
                'legacy_contract': str(self._cell(row, header_map, 'legacy_contract') or '').strip(),
                'meter_number': str(self._cell(row, header_map, 'meter_number') or '').strip(),
                'previous_balance': str(self._cell(row, header_map, 'previous_balance') or '').strip(),
                'current_balance': self.parse_float(self._cell(row, header_map, 'current_balance'), 'current_balance', row_number) or 0.0,
                'phase': self.parse_phase(self._cell(row, header_map, 'phase'), row_number),
                'is_private_transformer': self.parse_bool(self._cell(row, header_map, 'is_private_transformer'), False, 'is_private_transformer', row_number),
                'owner_reference': str(self._cell(row, header_map, 'owner_reference') or '').strip(),
                'meter_model_code': str(self._cell(row, header_map, 'meter_model_code') or '').strip(),
                'company_id': company_id, 'state': 'draft', 'source_row_number': row_number,
            }
            meter_reading = self.parse_float(self._cell(row, header_map, 'meter_reading'), 'meter_reading', row_number)
            opening_reading = self.parse_float(self._cell(row, header_map, 'opening_reading'), 'opening_reading', row_number)
            canonical_reading = opening_reading if opening_reading is not None else meter_reading
            if canonical_reading is not None:
                values.update(last_reading=canonical_reading, opening_reading=canonical_reading)
            existing = model.search([('company_id', '=', company_id), ('customer_number', '=', number)])
            if len(existing) > 1:
                raise self._error('AMBIGUOUS_CUSTOMER_IDENTITY', row_number, 'customer_number', number)
            record = existing or model.create(values)
            if existing:
                record.write(values)
            records |= record
        if records:
            records.action_map_codes(strict=False)
        return self._show_success_notification(len(records))

    def _import_feeders(self, sheet, header_row, header_map):
        model = self.env['utility.migration.feeder']
        records, seen = model, {}
        company_id = self.env.company.id
        for row_number, row in self._rows(sheet, header_row, header_map):
            code = str(self._cell(row, header_map, 'feeder_code') or '').strip()
            analytic = str(self._cell(row, header_map, 'legacy_analytic_id') or '').strip()
            identity = code or analytic
            if not identity:
                if not any(self._has_cell_value(v) for v in row):
                    continue
                raise self._error('MISSING_FEEDER_IDENTITY', row_number, 'feeder_code/legacy_analytic_id', identity)
            self._identity_guard(seen, identity, row_number, 'DUPLICATE_FEEDER_IDENTITY_IN_FILE')
            values = {
                'name': str(self._cell(row, header_map, 'feeder_name') or identity).strip(),
                'feeder_code': code, 'legacy_analytic_id': analytic,
                'feeder_name': str(self._cell(row, header_map, 'feeder_name') or identity).strip(),
                'legacy_region': str(self._cell(row, header_map, 'legacy_region') or '').strip(),
                'legacy_area': str(self._cell(row, header_map, 'legacy_area') or '').strip(),
                'is_active': self.parse_bool(self._cell(row, header_map, 'is_active'), True, 'is_active', row_number),
                'meter_number': str(self._cell(row, header_map, 'meter_number') or '').strip(),
                'meter_multiplier': self._parse_multiplier(self._cell(row, header_map, 'meter_multiplier'), row_number, 'meter_multiplier'),
                'is_calculation_cell': self.parse_bool(self._cell(row, header_map, 'is_calculation_cell'), True, 'is_calculation_cell', row_number),
                'description': str(self._cell(row, header_map, 'description') or '').strip(),
                'company_id': company_id, 'state': 'draft', 'source_row_number': row_number,
            }
            current = self.parse_float(self._cell(row, header_map, 'current_reading'), 'current_reading', row_number)
            opening = self.parse_float(self._cell(row, header_map, 'opening_reading'), 'opening_reading', row_number)
            if current is not None:
                values['current_reading'] = current
            elif opening is not None:
                values['opening_reading'] = opening
            existing = model.search([('company_id', '=', company_id), '|', ('feeder_code', '=', identity), ('legacy_analytic_id', '=', identity)])
            if len(existing) > 1:
                raise self._error('AMBIGUOUS_FEEDER_IDENTITY', row_number, 'identity', identity)
            record = existing or model.create(values)
            if existing:
                record.write(values)
            records |= record
        if records:
            records.action_map_codes(strict=False)
        return self._show_success_notification(len(records))

    def _import_transformers(self, sheet, header_row, header_map):
        model = self.env['utility.migration.transformer']
        records, seen = model, {}
        company_id = self.env.company.id
        for row_number, row in self._rows(sheet, header_row, header_map):
            reference = str(self._cell(row, header_map, 'reference') or '').strip()
            code = str(self._cell(row, header_map, 'transformer_code') or '').strip()
            analytic = str(self._cell(row, header_map, 'legacy_analytic_id') or '').strip()
            identity = reference or code or analytic
            if not identity:
                if not any(self._has_cell_value(v) for v in row):
                    continue
                raise self._error('MISSING_TRANSFORMER_IDENTITY', row_number, 'reference/transformer_code/legacy_analytic_id', identity)
            self._identity_guard(seen, identity, row_number, 'DUPLICATE_TRANSFORMER_IDENTITY_IN_FILE')
            values = {
                'name': str(self._cell(row, header_map, 'transformer_name') or identity).strip(),
                'reference': reference, 'transformer_code': code, 'legacy_analytic_id': analytic,
                'transformer_name': str(self._cell(row, header_map, 'transformer_name') or identity).strip(),
                'legacy_region': str(self._cell(row, header_map, 'legacy_region') or '').strip(),
                'legacy_area': str(self._cell(row, header_map, 'legacy_area') or '').strip(),
                'is_active': self.parse_bool(self._cell(row, header_map, 'is_active'), True, 'is_active', row_number),
                'meter_number': str(self._cell(row, header_map, 'meter_number') or '').strip(),
                'meter_multiplier': self._parse_multiplier(self._cell(row, header_map, 'meter_multiplier'), row_number, 'meter_multiplier'),
                'total_consumption': self.parse_float(self._cell(row, header_map, 'total_consumption'), 'total_consumption', row_number) or 0.0,
                'image_status': str(self._cell(row, header_map, 'image_status') or '').strip(),
                'cell_meter_number': str(self._cell(row, header_map, 'cell_meter_number') or '').strip(),
                'cell_meter_multiplier': self._parse_multiplier(self._cell(row, header_map, 'cell_meter_multiplier'), row_number, 'cell_meter_multiplier'),
                'description': str(self._cell(row, header_map, 'description') or '').strip(),
                'company_id': company_id, 'state': 'draft', 'source_row_number': row_number,
            }
            current = self.parse_float(self._cell(row, header_map, 'current_reading'), 'current_reading', row_number)
            opening = self.parse_float(self._cell(row, header_map, 'opening_reading'), 'opening_reading', row_number)
            if current is not None:
                values['current_reading'] = current
            elif opening is not None:
                values['opening_reading'] = opening
            existing = model.search([('company_id', '=', company_id), '|', ('reference', '=', identity), '|', ('transformer_code', '=', identity), ('legacy_analytic_id', '=', identity)])
            if len(existing) > 1:
                raise self._error('AMBIGUOUS_TRANSFORMER_IDENTITY', row_number, 'identity', identity)
            record = existing or model.create(values)
            if existing:
                record.write(values)
            records |= record
        if records:
            records.action_map_codes(strict=False)
        return self._show_success_notification(len(records))

    def _parse_multiplier(self, value, row, field):
        if not self._has_cell_value(value):
            return 1.0
        result = self.parse_float(value, field, row)
        if result is None or result <= 0:
            raise self._error('INVALID_METER_MULTIPLIER', row, field, value)
        return result

    def _show_success_notification(self, count):
        return {'type': 'ir.actions.act_window_close', 'tag': 'display_notification', 'params': {
            'title': _('تم رفع البيانات المبدئية بنجاح'),
            'message': _('تم رفع %s سجل إلى مرحلة التهيئة (Staging).') % count,
            'sticky': False, 'type': 'success', 'next': {'type': 'ir.actions.act_window_close'},
        }}
